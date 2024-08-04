import openpyxl
from django.conf import settings

class Excel:
    ALPHABETS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
    COLUMNS = list(ALPHABETS) + [f'A{i}' for i in ALPHABETS]
    
    def __init__(self, filename):
        self.filename = filename
        self.filepath = str(settings.TABLES_DIR / filename) + '.xlsx'
    
    def load(self):
        try:
            wb = openpyxl.load_workbook(self.filepath)
            ws = wb.active
            return wb, ws
        except Exception as e:
            print('ERROR:', e)
            exit()

    def setup(self, rows, columns):
        # Setup file.
        openpyxl.Workbook().save(self.filepath)
        wb, ws = self.load()
        
        # Setup rows.
        row_counter = 2
        for content in rows:
            ws['A' + str(row_counter)] = content
            row_counter += 1
        
        # Setup columns
        col_counter = 1
        for content in columns:
            ws[self.COLUMNS[col_counter] + '1'].value = content
            col_counter += 1

        wb.save(self.filepath)
    
    def write(self, solution):
        wb, ws = self.load()
        for k, v in solution.items():
            row = self.find_in_row(tuple(ws.rows), v['timeslot'].code)
            col = self.find_in_col(tuple(ws.columns), v['room'].code)
            cell = col + row
            ws[cell] = k.__str__()
        wb.save(self.filepath)
            
    def find_in_row(self, rows, value):
        """Find value in rows."""
        for n, row in enumerate(rows):
            for cell in row:
                if cell.value == None: continue
                if cell.value == value or value in cell.value:
                    return str(n + 1)
    
    def find_in_col(self, columns, value):
        for n, column in enumerate(columns):
            for cell in column:
                if cell.value == None: continue
                if cell.value == value or value in cell.value:
                    return self.COLUMNS[n]